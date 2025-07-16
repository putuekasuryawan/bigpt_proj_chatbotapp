from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.views.decorators.csrf import csrf_exempt
from .utils import send_bulk_template_message
from django.contrib.auth import logout
from django.shortcuts import redirect, render
from django.utils.timezone import make_aware
from bigptchatapp.models import *
from chatbotappbianco.models import *
from django.http import JsonResponse
from openpyxl import load_workbook
from django.conf import settings
from io import BytesIO
import pandas as pd
import random, string
import pdfplumber
import requests
import openai
import base64
import boto3
import pytz
import json
import os

openai.api_key = os.environ.get("OPENAI_KEY")

@login_required(login_url='signin')

def index(request):
    get_role = tbl_application_user_role.objects.filter(user_id=request.user).last().role_id
    granted_app = tbl_application_roles_setup.objects.filter(role_id=get_role).values_list('granted_app', flat=True)
    granted_app_endpoint = []
    for i in list(granted_app):
        app = tbl_application.objects.filter(id=i).last().endpoint
        granted_app_endpoint.insert(0, app)
    
    if request.build_absolute_uri() in granted_app_endpoint:
        userfullname = request.user.first_name + " " + request.user.last_name
        userinitial = request.user.first_name[0].capitalize() + request.user.last_name[0].capitalize()
        useremail = request.user.email
        userrole = tbl_application_roles.objects.filter(id=get_role).last().role_name
        application = tbl_application.objects.filter(endpoint=request.build_absolute_uri())
        this_app = application.last().pk
        usermenuid = tbl_application_roles_setup_menu.objects.filter(role_id=get_role).filter(app_id=this_app).order_by('-menu_queue').values_list('menu_id', flat=True)
        usermenu = []
        for n in list(usermenuid): 
            flag = tbl_application_roles_setup_menu.objects.filter(role_id=get_role).filter(app_id=this_app).filter(menu_id=n)
            obj = tbl_application_menu.objects.filter(id=n)
            if flag.last().status == 1:
                names = obj.last().menu_name
                icons = obj.last().menu_icon
                if obj.last().endpoint != "":
                    endpoints = request.build_absolute_uri() + str(this_app) + "/" + obj.last().endpoint
                else:
                    endpoints = request.build_absolute_uri() + obj.last().endpoint
                menu_html = '''
                                <li class="nk-menu-item"> 
                                    <a href="'''+ endpoints +'''" class="nk-menu-link">
                                        <span class="nk-menu-icon"><em class="icon ni '''+ icons +'''"></em></span>
                                        <span class="nk-menu-text">'''+ names +'''</span>
                                    </a>
                                </li>
                            '''
                usermenu.insert(0, menu_html)
                if obj.last().menu_name == "Setup":
                    setup_link = request.build_absolute_uri() + str(this_app) + "/" + obj.last().endpoint
        context = {"userfullname": userfullname, 
                   "userinitial": userinitial, 
                   "useremail": useremail, 
                   "userrole": userrole, 
                   "usermenus": usermenu,
                   "setup_link": setup_link}
        return render(request, str(application.last().app_name) + "/index.html", context)
    else:
        get_default_app = tbl_application_roles_setup.objects.filter(role_id=get_role).filter(default=1).last().granted_app
        default_endpoint = tbl_application.objects.filter(id=get_default_app).last().endpoint
        return redirect(default_endpoint)

def biancosetup(request, id):
    get_role = tbl_application_user_role.objects.filter(user_id=request.user).last().role_id
    granted_menu = tbl_application_roles_setup_menu.objects.filter(role_id=get_role).filter(app_id=int(id)).filter(status=1).order_by('-menu_queue').values_list('menu_id', flat=True)
    app = tbl_application.objects.filter(id=int(id))
    app_endpoint = app.last().endpoint
    granted_menu_endpoint = []
    for i in list(granted_menu):
        menu_endpoint = tbl_application_menu.objects.filter(id=i).last().endpoint
        if menu_endpoint != "":
            full_uri = app_endpoint + id + "/" + menu_endpoint
        else:
            full_uri = app_endpoint + menu_endpoint
        granted_menu_endpoint.insert(0, full_uri)    
    if request.build_absolute_uri() in granted_menu_endpoint:
        userfullname = request.user.first_name + " " + request.user.last_name
        userinitial = request.user.first_name[0].capitalize() + request.user.last_name[0].capitalize()
        useremail = request.user.email
        userrole = tbl_application_roles.objects.filter(id=get_role).last().role_name
        usermenu = []
        for n in list(granted_menu):
            flag = tbl_application_roles_setup_menu.objects.filter(role_id=get_role).filter(app_id=id).filter(menu_id=n)
            obj = tbl_application_menu.objects.filter(id=n)
            if flag.last().status == 1:
                names = obj.last().menu_name
                icons = obj.last().menu_icon
                menu_endpoint = tbl_application_menu.objects.filter(id=n).last().endpoint
                if menu_endpoint != "":
                    endpoints = app_endpoint + id + "/" + menu_endpoint
                else:
                    endpoints = app_endpoint + menu_endpoint
                menu_html = '''
                                <li class="nk-menu-item"> 
                                    <a href="'''+ endpoints +'''" class="nk-menu-link">
                                        <span class="nk-menu-icon"><em class="icon ni '''+ icons +'''"></em></span>
                                        <span class="nk-menu-text">'''+ names +'''</span>
                                    </a>
                                </li>
                            '''
                usermenu.insert(0, menu_html)
        prefix = tbl_twilioprefix.objects.using("chatbotdb").filter(owner = "bianco")
        suffix = tbl_twiliosuffix.objects.using("chatbotdb").filter(owner = "bianco")
        context = tbl_twiliocontext.objects.using("chatbotdb").filter(owner = "bianco")
        showprefix = ""
        showsuffix = ""
        showcontext = ""
        if context.count() > 0:
            showcontext = context.last().context
        if prefix.count() > 0:
            showprefix = prefix.last().prefix
        if suffix.count() > 0:
            showsuffix = suffix.last().suffix
        context = {"userfullname": userfullname, 
                   "userinitial": userinitial, 
                   "useremail": useremail, 
                   "userrole": userrole,
                   "usermenus": usermenu,
                   "showprefix": showprefix,
                   "showsuffix": showsuffix,
                   "showkonversi": showcontext}
        return render(request, str(app.last().app_name) + "/biancosetup.html", context)
    else:
        return redirect(app_endpoint)

def biancotemplates(request, id):
    get_role = tbl_application_user_role.objects.filter(user_id=request.user).last().role_id
    granted_menu = tbl_application_roles_setup_menu.objects.filter(role_id=get_role).filter(app_id=int(id)).filter(status=1).order_by('-menu_queue').values_list('menu_id', flat=True)
    app = tbl_application.objects.filter(id=int(id))
    app_endpoint = app.last().endpoint
    granted_menu_endpoint = []
    for i in list(granted_menu):
        menu_endpoint = tbl_application_menu.objects.filter(id=i).last().endpoint
        if menu_endpoint != "":
            full_uri = app_endpoint + id + "/" + menu_endpoint
        else:
            full_uri = app_endpoint + menu_endpoint
        granted_menu_endpoint.insert(0, full_uri)    
    if request.build_absolute_uri() in granted_menu_endpoint:
        userfullname = request.user.first_name + " " + request.user.last_name
        userinitial = request.user.first_name[0].capitalize() + request.user.last_name[0].capitalize()
        useremail = request.user.email
        userrole = tbl_application_roles.objects.filter(id=get_role).last().role_name
        usermenu = []
        for n in list(granted_menu):
            flag = tbl_application_roles_setup_menu.objects.filter(role_id=get_role).filter(app_id=id).filter(menu_id=n)
            obj = tbl_application_menu.objects.filter(id=n)
            if flag.last().status == 1:
                names = obj.last().menu_name
                icons = obj.last().menu_icon
                menu_endpoint = tbl_application_menu.objects.filter(id=n).last().endpoint
                if menu_endpoint != "":
                    endpoints = app_endpoint + id + "/" + menu_endpoint
                else:
                    endpoints = app_endpoint + menu_endpoint
                menu_html = '''
                                <li class="nk-menu-item"> 
                                    <a href="'''+ endpoints +'''" class="nk-menu-link">
                                        <span class="nk-menu-icon"><em class="icon ni '''+ icons +'''"></em></span>
                                        <span class="nk-menu-text">'''+ names +'''</span>
                                    </a>
                                </li>
                            '''
                usermenu.insert(0, menu_html)
        prefix = tbl_twilioprefix.objects.using("chatbotdb").filter(owner = "bianco")
        suffix = tbl_twiliosuffix.objects.using("chatbotdb").filter(owner = "bianco")
        context = tbl_twiliocontext.objects.using("chatbotdb").filter(owner = "bianco")
        showprefix = ""
        showsuffix = ""
        showcontext = ""
        if context.count() > 0:
            showcontext = context.last().context
        if prefix.count() > 0:
            showprefix = prefix.last().prefix
        if suffix.count() > 0:
            showsuffix = suffix.last().suffix
        context = {"userfullname": userfullname, 
                   "userinitial": userinitial, 
                   "useremail": useremail, 
                   "userrole": userrole,
                   "usermenus": usermenu,
                   "showprefix": showprefix,
                   "showsuffix": showsuffix,
                   "showkonversi": showcontext}
        return render(request, str(app.last().app_name) + "/biancotemplate.html", context)
    else:
        return redirect(app_endpoint)

@csrf_exempt
def instruksicpsaved(request):
    data = json.loads(request.body)
    state = ""
    res = ""
    try:
        prefix = tbl_twilioprefix.objects.using("chatbotdb").filter(owner="bianco")
        if prefix.count() > 0:
            prefix.update(prefix = data["profile"])
            state = "success"
            res = "Save data profile to database completed"
        else:
            prefix.create( 
                owner = "bianco",
                prefix = data["profile"]
            )
            state = "success"
            res = "Save data profile to database completed"
    except Exception as e:
        state = "failed"
        res = "Failed to save data to database, error : " + str(e)
    return JsonResponse({"result": state, "message" : res})

@csrf_exempt
def instruksisuffixsaved(request):
    data = json.loads(request.body)
    state = ""
    res = ""
    try:
        suffix = tbl_twiliosuffix.objects.using("chatbotdb").filter(owner="bianco")
        if suffix.count() > 0:
            suffix.update(suffix = data["suffix"])
            state = "success"
            res = "Save data suffix to database completed"
        else:
            suffix.create( 
                owner = "bianco",
                suffix = data["suffix"]
            )
            state = "success"
            res = "Save data suffix to database completed"
    except Exception as e:
        state = "failed"
        res = "Failed to save data to database, error : " + str(e)
    return JsonResponse({"result": state, "message" : res})

def signout(request):
    logout(request)
    return redirect("signin")

def detect_file_type(ext):
    if ext == '.pdf':
        return 'pdf'
    elif ext in ['.xls', '.xlsx']:
        return 'excel'
    elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp']:
        return 'images'
    else:
        return 'unknown'

@csrf_exempt
def fileuploadbianco(request):
    folder_name = ""
    size = 6
    res = "".join(random.choices(string.ascii_uppercase + string.digits, k=size))
    file_path = ""
    minio_url = ""
    request_file = request.FILES.get("file")
    if not request_file:
        return JsonResponse({"result": "failed", "message": "No file provided"})
    file_name = request_file.name
    file_ext = os.path.splitext(file_name)[1].lower()
    file_type = detect_file_type(file_ext)
    if file_type == "unknown":
        state = "failed"
        res = "gagal upload file ke MinIO storage, file extension not supported"
    else:
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        if file_type == "pdf":
            folder_name = f"bianco/{file_type}/"
        elif file_type == "excel":
            folder_name = f"bianco/{file_type}/"
        elif file_type == "images":
            folder_name = f"bianco/{file_type}/"
        file_path = f"{folder_name}BIANCO-{res}-{file_name}"
        minio_url = f"{settings.AWS_S3_ENDPOINT_URL}/{bucket_name}/{file_path}"
        print("file_type:", file_type)
        print("folder_name:", folder_name)
        print("file_path:", file_path)
        print("minio_url:", minio_url)
        try:
            default_storage.save(file_path, request_file)
            state = "success"
            res = "upload file ke MinIO storage berhasil"
        except Exception as e:
            print("Exception:", e)
            state = "failed"
            res = "gagal upload file ke MinIO storage, " + str(e)
    return JsonResponse({"result": state, "message": res, "type": file_type, "generatedpath": file_path, "fullpath": minio_url})

def get_file_from_minio(bucket_name, file_path):
    s3 = boto3.client(
        "s3",
        endpoint_url=settings.AWS_S3_ENDPOINT_URL,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        region_name=settings.AWS_S3_REGION_NAME
    )
    response = s3.get_object(Bucket=bucket_name, Key=file_path)
    return response["Body"].read()

def extract_text_from_pdf(file_bytes):
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        text = "\n".join(page.extract_text() for page in pdf.pages if page.extract_text())
    return text

def extract_text_from_excel(file_bytes):
    excel_file = BytesIO(file_bytes)
    wb = load_workbook(excel_file, data_only=True)
    all_text = []
    for sheet in wb.worksheets:
        merged_ranges = sheet.merged_cells.ranges
        sheet_text = [f"=== Sheet: {sheet.title} ==="]
        for row in sheet.iter_rows():
            row_values = []
            for cell in row:
                value = cell.value
                if value is None:
                    for merge_range in merged_ranges:
                        if cell.coordinate in merge_range:
                            master_cell = sheet[merge_range.coord.split(":")[0]]
                            value = master_cell.value
                            break
                value_str = str(value).strip() if value is not None else ""
                row_values.append(value_str)
            row_text = " | ".join(row_values).rstrip(" |")
            sheet_text.append(row_text)
        all_text.append("\n".join(sheet_text))
    return "\n\n".join(all_text)

def ask_gpt_to_convert_to_json(raw_text: str) -> str:
    prompt = (
        "Tolong konversikan teks berikut menjadi JSON yang valid, dengan format yang sesuai:\n\n"
        f"{raw_text}\n\n"
        "Berikan hanya JSON (Expecting property name enclosed in double quotes) sebagai output tanpa penjelasan, tanpa kode markdown apapun."
    )
    response = openai.chat.completions.create(
        model="gpt-4.1-2025-04-14",
        messages=[
            {"role": "system", "content": "Kamu adalah asisten yang ahli dalam parsing teks menjadi JSON."},
            {"role": "user", "content": prompt},
        ],
        max_tokens=32000,
        temperature=0.0,
        n=1,
        stop=None
    )
    return response.choices[0].message.content

@csrf_exempt
def konversi_started_pdf(request):
    data = json.loads(request.body)
    json_result = ""
    state = ""
    res = ""
    try:
        bucket = settings.AWS_STORAGE_BUCKET_NAME
        path = data["path"]
        file_bytes = get_file_from_minio(bucket, path)
        raw_teks = extract_text_from_pdf(file_bytes)
        json_result = ask_gpt_to_convert_to_json(raw_teks)
        parsed_json = json.loads(json_result)
        tbl_twiliocontext.objects.using("chatbotdb").create(owner = "bianco", context = json_result)
        state = "success"
        res = "proses konversi berhasil dijalankan"
    except Exception as e:
        state = "failed"
        res = "gagal dalam proses konversi, error : " + str(e)
    return JsonResponse({"result": state, "message": res})

@csrf_exempt
def konversi_started_excel(request):
    data = json.loads(request.body)
    json_result = ""
    state = ""
    res = ""
    try:
        bucket = settings.AWS_STORAGE_BUCKET_NAME
        path = data["path"]
        file_bytes = get_file_from_minio(bucket, path)
        raw_teks = extract_text_from_excel(file_bytes)
        json_result = ask_gpt_to_convert_to_json(raw_teks)
        parsed_json = json.loads(json_result)
        tbl_twiliocontext.objects.using("chatbotdb").create(owner = "bianco", context = json_result)
        state = "success"
        res = "proses konversi berhasil dijalankan"
    except Exception as e:
        state = "failed"
        res = "gagal dalam proses konversi, error : " + str(e)
    return JsonResponse({"result": state, "message": res})

def format_date(date_str):
    if date_str:
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%SZ")
            dt = make_aware(dt, pytz.utc)
            return dt
        except ValueError:
            return make_aware(datetime.now(), pytz.utc)
    else:
        return make_aware(datetime.now(), pytz.utc)

@csrf_exempt
def contenttypetextapi(request):
    TWILIO_ACCOUNT_SID = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().account_sid
    TWILIO_AUTH_TOKEN = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().auth_token
    TWILIO_API_URL = "https://content.twilio.com/v1/Content"
    state = ""
    res = ""
    data = json.loads(request.body) 
    variables = {f"{i}": f"sample_var{i}" for i in range(1, int(data["numbers"]) + 1)}
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
    }
    payload = json.dumps({
        "friendly_name": data["fname"],
        "language":  data["lang"],
        "variables": variables,
        "types": {
            "twilio/text": {
                "body": data["body"]
            }
        }
    })
    try:
        response = requests.request("POST", TWILIO_API_URL, headers=headers, data=payload)
        if response.status_code == 201:
            datak = response.json()
            content_sid = datak.get("sid")
            body_value = datak.get("types", {}).get("twilio/text", {}).get("body", "")
            approval_create_url = datak.get("links", {}).get("approval_create", "")
            date_created = format_date(datak.get("date_created")) 
            
            tbl_chattemplate.objects.using("chatbotdb").create(user = request.user, type = "text", content_sid = content_sid, friendly_name = datak.get("friendly_name"), language = datak.get("language"), numbers = int(data["numbers"]), body = body_value, submit_status = "Not Submitted", created = date_created)
            
            headers2 = {
                "Content-Type": "application/json",
                "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
            }
            payload2 = json.dumps({
                "name": datak.get("friendly_name"),
                "category":  "MARKETING"
            })
            response1 = requests.request("POST", approval_create_url, headers=headers2, data=payload2)
            if response1.status_code == 201:
                dataku = response1.json()
                
                objs = tbl_chattemplate.objects.using("chatbotdb").filter(content_sid = content_sid)
                if objs.count() > 0:
                    objs.update(submit_status = dataku.get("status"))
                    state = "success"
                    res = "Proses submit ke Meta berhasil dijalankan"
                else:
                    state = "success"
                    res = "Proses submit ke Meta berhasil tapi gagal dalam update data ke database"
            else:
                state = "failed"
                print(str(response1.text))
                res = "Proses submit ke Meta gagal dijalankan tapi data berhasil masuk ke Twilio Console, error code: " + str(response1.status_code) + " (" + str(response1.text) + ")"
        else:
            state = "failed"
            res = "Proses submit gagal dengan status code: " + str(response.status_code) + " (" + str(response.text) + ")"
    except Exception as e:
        state = "failed"
        res = "Proses submit gagal, error : " + str(e)
    return JsonResponse({"result": state, "message" : res})

@csrf_exempt
def contenttypemediaapi(request):
    TWILIO_ACCOUNT_SID = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().account_sid
    TWILIO_AUTH_TOKEN = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().auth_token
    TWILIO_API_URL = "https://content.twilio.com/v1/Content"
    state = ""
    res = ""
    data = json.loads(request.body) 
    variables = {f"{i}": f"sample_var{i}" for i in range(1, int(data["numbers"]) + 1)}
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
    }
    payload = json.dumps({
        "friendly_name": data["fname"],
        "language":  data["lang"],
        "variables": variables,
        "types": {
            "twilio/media": {
                "body": data["body"],
                "media": [data["mediaurl"]]
            }
        }
    })
    try:
        response = requests.request("POST", TWILIO_API_URL, headers=headers, data=payload)
        if response.status_code == 201:
            datak = response.json()
            content_sid = datak.get("sid")
            body_value = datak.get("types", {}).get("twilio/media", {}).get("body", "")
            url_value = datak.get("types", {}).get("twilio/media", {}).get("media", "")
            approval_create_url = datak.get("links", {}).get("approval_create", "")
            date_created = format_date(datak.get("date_created")) 
            
            tbl_chattemplate.objects.using("chatbotdb").create(user = request.user, type = "media", content_sid = content_sid, friendly_name = datak.get("friendly_name"), language = datak.get("language"), numbers = int(data["numbers"]), body = body_value, media_url = url_value, submit_status = "Not Submitted", created = date_created)
            
            headers2 = {
                "Content-Type": "application/json",
                "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
            }
            payload2 = json.dumps({
                "name": datak.get("friendly_name"),
                "category":  "MARKETING"
            })
            response1 = requests.request("POST", approval_create_url, headers=headers2, data=payload2)
            if response1.status_code == 201:
                dataku = response1.json()
                
                objs = tbl_chattemplate.objects.using("chatbotdb").filter(content_sid = content_sid)
                if objs.count() > 0:
                    objs.update(submit_status = dataku.get("status"))
                    state = "success"
                    res = "Proses submit ke Meta berhasil dijalankan"
                else:
                    state = "success"
                    res = "Proses submit ke Meta berhasil tapi gagal dalam update data ke database"
            else:
                state = "failed"
                print(str(response1.text))
                res = "Proses submit ke Meta gagal dijalankan tapi data berhasil masuk ke Twilio Console, error code: " + str(response1.status_code) + " (" + str(response1.text) + ")"
        else:
            state = "failed"
            res = "Proses submit gagal dengan status code: " + str(response.status_code) + " (" + str(response.text) + ")"
    except Exception as e:
        state = "failed"
        res = "Proses submit gagal, error : " + str(e)
    return JsonResponse({"result": state, "message" : res})

@csrf_exempt
def get_template_status(request):
    TWILIO_ACCOUNT_SID = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().account_sid
    TWILIO_AUTH_TOKEN = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().auth_token
    TWILIO_API_URL = "https://content.twilio.com/v1/Content"
    result_data = []
    if request.method == "GET" and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        templates = tbl_chattemplate.objects.using("chatbotdb").all()
        headers = {
                "Content-Type": "application/json",
                "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
            }
        for s in templates:
            url_fetch = f"{TWILIO_API_URL}/{s.content_sid}/ApprovalRequests"
            try:
                responses = requests.request("GET", url_fetch, headers=headers)
                if responses.status_code == 200:
                    data = responses.json()
                    result_data.append({
                        "friendly_name": s.friendly_name,
                        "type": s.type,
                        "content_sid": s.content_sid,
                        "submit_status": data["whatsapp"]["status"],
                        "created": s.created
                    })
                else:
                    print(f"Failed to fetch from Twilio: {responses.status_code} - {responses.text}")
            except Exception as e:
                print("Exception occurred:", str(e))
                result_data = [{}]
        return JsonResponse({"data": result_data})
    
@csrf_exempt
def delete_template(request, sid):
    TWILIO_ACCOUNT_SID = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().account_sid
    TWILIO_AUTH_TOKEN = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().auth_token
    TWILIO_API_URL = "https://content.twilio.com/v1/Content"
    if request.method == "POST":
        final_url = f"{TWILIO_API_URL}/{sid}"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
        }
        try:
            responsess = requests.request("DELETE", final_url, headers=headers)
            if responsess.status_code == 204:
                template = tbl_chattemplate.objects.using("chatbotdb").get(content_sid=sid)
                template.delete()
            return JsonResponse({"status": "success", "message": "Template deleted successfully."})
        except tbl_chattemplate.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Template not found."}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    else:
        return JsonResponse({"status": "error", "message": "Invalid request method."}, status=405)

@csrf_exempt
def get_content_template(request, sid):
    TWILIO_ACCOUNT_SID = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().account_sid
    TWILIO_AUTH_TOKEN = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().auth_token
    TWILIO_API_URL = "https://content.twilio.com/v1/Content"
    result_data = []
    if request.method == "GET" and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        url_fetch = f"{TWILIO_API_URL}/{sid}"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
        }
        try:
            responsesss = requests.request("GET", url_fetch, headers=headers)
            if responsesss.status_code == 200:
                data = responsesss.json()
                if "twilio/text" in data["types"]:
                    result_data.append({
                        "friendly_name": data["friendly_name"],
                        "type": "text",
                        "content_sid": data["sid"],
                        "body": data["types"]["twilio/text"]["body"],
                        "media": ""
                    })
                elif "twilio/media" in data["types"]:
                    result_data.append({
                        "friendly_name": data["friendly_name"],
                        "type": "media",
                        "content_sid": data["sid"],
                        "body": data["types"]["twilio/media"]["body"],
                        "media": data["types"]["twilio/media"]["media"][0]
                    })
        except Exception as e:
                print("Exception occurred:", str(e))
                result_data = [{}]
    return JsonResponse({"data": result_data})       

@csrf_exempt
def get_template_approved(request):
    TWILIO_ACCOUNT_SID = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().account_sid
    TWILIO_AUTH_TOKEN = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().auth_token
    TWILIO_API_URL = "https://content.twilio.com/v1/Content"
    result_data = []
    if request.method == "GET" and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        templates = tbl_chattemplate.objects.using("chatbotdb").order_by('-created')
        headers = {
                "Content-Type": "application/json",
                "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
            }
        for s in templates:
            url_fetch = f"{TWILIO_API_URL}/{s.content_sid}/ApprovalRequests"
            try:
                responses = requests.request("GET", url_fetch, headers=headers)
                if responses.status_code == 200:
                    data = responses.json()
                    if "approved" in data["whatsapp"]["status"]:
                        result_data.append({
                            "friendly_name": s.friendly_name,
                            "type": s.type,
                            "content_sid": s.content_sid,
                            "submit_status": data["whatsapp"]["status"],
                            "created": s.created
                        })
                else:
                    print(f"Failed to fetch from Twilio: {responses.status_code} - {responses.text}")
            except Exception as e:
                print("Exception occurred:", str(e))
                result_data = [{}]
        return JsonResponse({"data": result_data})

@csrf_exempt
def get_template_approved_detail(request, content_sid):
    TWILIO_ACCOUNT_SID = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().account_sid
    TWILIO_AUTH_TOKEN = tbl_twiliocredentials.objects.using("chatbotdb").filter(type="production").last().auth_token
    TWILIO_API_URL = "https://content.twilio.com/v1/Content"
    template = tbl_chattemplate.objects.using("chatbotdb").filter(content_sid=content_sid)
    if not template:
        return JsonResponse({"error": "Not found"}, status=404)

    result_data = []
    if request.method == "GET" and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        url_fetch = f"{TWILIO_API_URL}/{content_sid}"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Basic {base64.b64encode(f'{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}'.encode()).decode()}"
        }
        try:
            responsesss = requests.request("GET", url_fetch, headers=headers)
            if responsesss.status_code == 200:
                data = responsesss.json()
                if "twilio/text" in data["types"]:
                    result_data.append({
                        "friendly_name": data["friendly_name"],
                        "type": "text",
                        "content_sid": data["sid"],
                        "body": data["types"]["twilio/text"]["body"],
                        "media": "",
                        "varsnumber": template.last().numbers
                    })
                elif "twilio/media" in data["types"]:
                    result_data.append({
                        "friendly_name": data["friendly_name"],
                        "type": "media",
                        "content_sid": data["sid"],
                        "body": data["types"]["twilio/media"]["body"],
                        "media": data["types"]["twilio/media"]["media"][0],
                        "varsnumber": template.last().numbers
                    })
        except Exception as e:
                print("Exception occurred:", str(e))
                result_data = [{}]
    return JsonResponse({"data": result_data})

@csrf_exempt
def templatekonversi(request):
    state = ""
    res = ""
    data = json.loads(request.body)
    file_path = data["fullpath"]
    print(file_path)

    response = requests.get(file_path)
    response.raise_for_status()
    df = pd.read_excel(BytesIO(response.content))
    if 'phone_number' not in df.columns:
        state = "failed"
        res = "Kolom 'phone_number' wajib ada."
    else:
        recipients = []
        var_cols = [col for col in df.columns if col != 'phone_number']
        for _, row in df.iterrows():
            entry = {"number": str(row['phone_number'])}
            for i, col in enumerate(var_cols):
                value = row[col]
                if pd.notna(value):
                    entry[str(i + 1)] = str(value)
            recipients.append(entry)
        print(recipients)
        state = "success"
        res = "Proses konversi berjalan lancar"
    return JsonResponse({"result": state, "message" : res, "output" : recipients})

@csrf_exempt
def sendblasttemplate(request):
    state = ""
    res = ""
    data = json.loads(request.body)
    sid = data["sid"]
    recipients = data["recipients"]
    try:
        send_bulk_template_message(sid, recipients)
        state = "success"
        res = "Blast whatsapp template berhasil dikirimkan"
    except Exception as e:
        state = "failed"
        res = "Blast whatsapp template gagal dikirimkan. error: " + str(e)
        
    return JsonResponse({"result": state, "message" : res}) 
